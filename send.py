import requests
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import datetime
from email.mime.text import MIMEText
import smtplib
import random


# 从xlsx文件中读取数据
def get_info_from_xlsx():
    username = []
    password = []
    userEmail = []
    wb = load_workbook(filename='config.xlsx')
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    row = wb[sheetnames[0]].max_row
    for rowNum in range(1, row + 1):
        # print(sheet.cell(row=rowNum, column=1).value)
        username.append(sheet.cell(row=rowNum, column=1).value)
        # print(sheet.cell(row=rowNum, column=2).value)
        password.append(sheet.cell(row=rowNum, column=2).value)
        # print(sheet.cell(row=rowNum, column=3).value)
        userEmail.append(sheet.cell(row=rowNum, column=3).value)
    sum = len(username)
    return username, password, userEmail, sum


# 从txt文件中读取数据
def get_info_from_txt():
    username = []
    password = []
    userEmail = []
    name = []
    sum = 0
    configFile = open('config.txt', 'r', encoding='utf-8')
    list = configFile.readlines()
    # print(list)
    for lists in list:
        lists = lists.strip()
        lists = lists.strip('\n')
        lists = lists.strip(',')
        lists = lists.split(',')
        # print(lists[0])
        username.append(lists[0])
        # print(lists[1])
        password.append(lists[1])
        # print(lists[2])
        userEmail.append(lists[2])
        # print(lists[3])
        name.append(lists[3])
        sum += 1
    # print('sum: ', sum)
    return username, password, userEmail, name, sum


# 登录，为后续做准备
def login(s, headers, username, password):
    login_url = 'http://yiqing.ctgu.edu.cn/wx/index/loginSubmit.do'
    data = {
        'username': '',
        'password': '',
    }
    data['username'] = username
    data['password'] = password
    r = s.post(url=login_url, headers=headers, data=data)
    # 增加等待时间，让数据接收完毕
    time.sleep(5)
    return r.text


# 获取要带有data的html
def get_student_info(s, headers):
    student_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/toApply.do'
    r = s.get(url=student_info_url, headers=headers)
    # print(r.text)
    # 增加等待时间，让数据接收完毕
    time.sleep(20)
    return r.text


# 解析html,获得data数据
def student_info_parse(html):
    bs = BeautifulSoup(html, 'lxml')
    data = {
        'ttoken': bs.find(attrs={'name': 'ttoken'})['value'],
        # 省份
        'province': bs.find(attrs={'name': 'province'})['value'],
        # 城市
        'city': bs.find(attrs={'name': 'city'})['value'],
        # 县区
        'district': bs.find(attrs={'name': 'district'})['value'],
        # 地区代码：身份证号前六位
        'adcode': bs.find(attrs={'name': 'adcode'})['value'],
        'longitude': bs.find(attrs={'name': 'longitude'})['value'],
        'latitude': bs.find(attrs={'name': 'latitude'})['value'],
        # 是否确诊新型肺炎
        'sfqz': bs.find(attrs={'name': 'sfqz'})['value'],
        # 是否疑似感染
        'sfys': bs.find(attrs={'name': 'sfys'})['value'],
        'sfzy': bs.find(attrs={'name': 'sfzy'})['value'],
        # 是否隔离
        'sfgl': bs.find(attrs={'name': 'sfgl'})['value'],
        'status': bs.find(attrs={'name': 'status'})['value'],
        'sfgr': bs.find(attrs={'name': 'sfgr'})['value'],
        'szdz': bs.find(attrs={'name': 'szdz'})['value'],
        # 手机号
        'sjh': bs.find(attrs={'name': 'sjh'})['value'],
        # 紧急联系人姓名
        'lxrxm': bs.find(attrs={'name': 'lxrxm'})['value'],
        # 紧急联系人手机号
        'lxrsjh': bs.find(attrs={'name': 'lxrsjh'})['value'],
        # 是否发热
        'sffr': '否',
        'sffy': '否',
        'sfgr': '否',
        'qzglsj': '',
        'qzgldd': '',
        'glyy': '',
        'mqzz': '',
        # 是否返校
        'sffx': '否',
        # 其它
        'qt': '',
    }
    # print(data)
    return data


# 向服务器post数据
def sent_info(s, headers, data):
    sent_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/saveApply.do'
    r = s.post(url=sent_info_url, headers=headers, data=data)
    # print(r.text)
    # 增加等待时间，让数据接收完毕
    time.sleep(20)
    print(r.status_code)


# 发邮件通知结果
def send_rusult(text, userEmail):
    # nowtime = datetime.datetime.now().strftime('%m-%d')
    msg_from = '2316453754@qq.com'  # 发送邮箱
    passwd = 'yrtorlthsiwuebgh'  # 密码
    msg_to = userEmail  # 目的邮箱

    subject = '安全上报结果'
    # content = text + '\n' + nowtime
    content = text
    msg = MIMEText(content)
    msg['Subject'] = subject
    msg['From'] = msg_from
    msg['To'] = msg_to

    # print(text)

    s = smtplib.SMTP_SSL("smtp.qq.com", 465)
    s.login(msg_from, passwd)
    s.sendmail(msg_from, msg_to, msg.as_string())
    print("邮件发送成功")
    s.quit()


def main():
    username, password, userEmail, name, sum = get_info_from_txt()
    # username, password, userEmail, sum = get_info_from_xlsx()
    print('---------------------------')
    print('账号密码读取成功，共', sum, '人')
    print('---------------------------\n')
    finished = 0
    reported = 0
    # 发送给用户的邮件信息
    userString = []
    # 发送给admin的邮件信息
    adminString = []

    adminString.append('---------------------------')
    adminString.append('账号密码读取成功，共' + str(sum) + '人')
    adminString.append('---------------------------' + '\n')
    try:
        for i in range(sum):
            print('')
            headers = {
                'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Mobile Safari/537.36',
                'Connection': 'close',
            }
            s = requests.session()
            s.keep_alive = False
            # print(s)
            # print(headers)
            state = login(s, headers, username[i], password[i])
            if state == 'success':
                print('用户', name[i], username[i], '登录成功')
                # 记录用户登录成功，添加到userString和adminString
                text = '用户 ' + str(name[i]) + str(username[i]) + '登录成功'
                userString.append(text)
                adminString.append(text)
            else:
                print('用户', name[i], username[i], '密码错误，登录失败')
                # 记录用户登录失败，添加到userString和adminString
                text = '用户 ' + str(name[i]) + str(username[i]) + '密码错误，登录失败'
                userString.append(text)
                adminString.append(text)
                # 将登录失败信息 邮件发送给用户
                send_rusult('\n'.join(userString), userEmail[i])
                continue
            html = get_student_info(s, headers)
            try:
                data = student_info_parse(html)
                sent_info(s, headers, data)
                # 记录用户上报成功，添加到userString和adminString
                text = '用户 ' + name[i] + '-' + str(username[i]) + ' 上报成功'
                print(text)
                userString.append(text)
                adminString.append(text + '\n')
                # print(userEmail[i])
                # send_rusult(text, userEmail[i])
                # 将用户上报成功信息，邮件发送给用户
                send_rusult('\n'.join(userString), userEmail[i])
            except:
                text = '用户 ' + name[i] + '-' + str(username[i]) + ' 今日已上报'
                print(text)
                adminString.append(text + '\n')
                reported += 1
                # print(userEmail[i])
                # send_rusult(text, userEmail[i])
            finished += 1
            # 增大等待间隔，github访问速度慢
            time.sleep(random.randint(10, 40))
            # time.sleep(random.randint(1, 10))
    finally:
        text = '应报:' + str(sum) + ' 本次上报:' + str(finished) + ' 今日已上报:' + str(reported)
        adminString.append('---------------------------')
        adminString.append(text)
        adminString.append('---------------------------')
        # print('\n'.join(adminString))

        print('---------------------------')
        print(text)
        # send_rusult(text, 'yizhaosan@qq.com')
        send_rusult('\n'.join(adminString), 'yizhaosan@qq.com')
        print('---------------------------\n')


if __name__ == "__main__":
    main()
